import xlrd
from config.pathdata import log_file_path
from config.pathdata import case_path
from openpyxl import load_workbook


class Case_operation:
    # 用例读取写入工具

    def get_case(self, filename, sheetname):
        wb = xlrd.open_workbook(filename=filename)#打开excl文件
        sheet = wb.sheet_by_name(sheetname)#按照表名打开对应工作表工具
        list1 = []
        case = []
        key = sheet.row_values(0) #获取第一行数据
        for i in range(1, sheet.nrows):
            d = []
            res = sheet.row_values(i)#遍历取得除第一行外所有行数
            for a in res:
                d.append(a)#将每一行的值插入空列表d中
            list1.append(d)
            for s in list1:
                r = dict(zip(key, s))#循环合并excle第一行和每一行的值
            case.append(r)
        return case

    def write_result(self, res, xulie):  # 测试结果写入
        wb = open(file=log_file_path, mode='r', encoding='utf-8').readlines() # 打开log文件并读取log文件
        wc = load_workbook(case_path)  # 打开excle文件
        data = []
        l = []
        for i in wb:
            data.append(i)  # 将文本文件的值存储至列表中
        for a in data:
            if res in a:
                ws = wc.active
                sheet = wc.worksheets[xulie]
                l.append(a)
        for i in range(len(l)):
            sheet.cell(i + 2, 8).value = l[i]
        wc.save(case_path)
